"""
Serviço para processar e extrair conteúdo de documentos.
Suporta PDF, DOCX, TXT, CSV, XLSX.
"""

import io
from pathlib import Path
from typing import List, Dict, Any
import logging

logger = logging.getLogger(__name__)


class DocumentProcessor:
    """Processa diferentes tipos de documentos e extrai texto."""

    def __init__(self):
        """Inicializa processador de documentos."""
        pass

    async def process_file(
        self,
        file_content: bytes,
        filename: str,
        chunk_size: int = 1000
    ) -> List[Dict[str, Any]]:
        """
        Processa arquivo e retorna lista de chunks de texto.

        Args:
            file_content: Conteúdo do arquivo em bytes
            filename: Nome do arquivo (para determinar tipo)
            chunk_size: Tamanho máximo de cada chunk em caracteres

        Returns:
            Lista de dicts com {"content": str, "chunk_index": int, "metadata": dict}

        Raises:
            ValueError: Se tipo de arquivo não suportado
        """
        file_extension = Path(filename).suffix.lower()

        # Determina método de processamento baseado na extensão
        processors = {
            '.pdf': self._process_pdf,
            '.txt': self._process_txt,
            '.docx': self._process_docx,
            '.doc': self._process_docx,
            '.csv': self._process_csv,
            '.xlsx': self._process_excel,
            '.xls': self._process_excel,
        }

        processor = processors.get(file_extension)
        if not processor:
            raise ValueError(f"Unsupported file type: {file_extension}")

        try:
            logger.info(f"Processing file: {filename} ({file_extension})")
            text = await processor(file_content)

            # Divide texto em chunks
            chunks = self._chunk_text(text, chunk_size)

            # Formata resultado
            result = []
            for i, chunk in enumerate(chunks):
                result.append({
                    "content": chunk,
                    "chunk_index": i,
                    "metadata": {
                        "source_file": filename,
                        "file_type": file_extension,
                        "total_chunks": len(chunks)
                    }
                })

            logger.info(f"Processed {filename}: {len(chunks)} chunks created")
            return result

        except Exception as e:
            logger.error(f"Failed to process file {filename}: {e}")
            raise

    async def _process_pdf(self, content: bytes) -> str:
        """
        Extrai texto de PDF.

        Args:
            content: Conteúdo do PDF

        Returns:
            Texto extraído
        """
        try:
            import PyPDF2

            pdf_file = io.BytesIO(content)
            pdf_reader = PyPDF2.PdfReader(pdf_file)

            text_parts = []
            for page_num, page in enumerate(pdf_reader.pages):
                text = page.extract_text()
                if text.strip():
                    text_parts.append(f"[Página {page_num + 1}]\n{text}")

            full_text = "\n\n".join(text_parts)
            logger.debug(f"Extracted {len(full_text)} chars from PDF")
            return full_text

        except Exception as e:
            logger.error(f"PDF processing error: {e}")
            raise ValueError(f"Failed to process PDF: {e}")

    async def _process_txt(self, content: bytes) -> str:
        """
        Processa arquivo TXT.

        Args:
            content: Conteúdo do arquivo

        Returns:
            Texto decodificado
        """
        try:
            # Tenta diferentes encodings
            encodings = ['utf-8', 'latin-1', 'cp1252']

            for encoding in encodings:
                try:
                    text = content.decode(encoding)
                    logger.debug(f"Decoded TXT with {encoding}")
                    return text
                except UnicodeDecodeError:
                    continue

            raise ValueError("Could not decode text file with supported encodings")

        except Exception as e:
            logger.error(f"TXT processing error: {e}")
            raise

    async def _process_docx(self, content: bytes) -> str:
        """
        Extrai texto de DOCX.

        Args:
            content: Conteúdo do DOCX

        Returns:
            Texto extraído
        """
        try:
            import docx

            doc_file = io.BytesIO(content)
            doc = docx.Document(doc_file)

            text_parts = []
            for para in doc.paragraphs:
                if para.text.strip():
                    text_parts.append(para.text)

            full_text = "\n\n".join(text_parts)
            logger.debug(f"Extracted {len(full_text)} chars from DOCX")
            return full_text

        except Exception as e:
            logger.error(f"DOCX processing error: {e}")
            raise ValueError(f"Failed to process DOCX: {e}")

    async def _process_csv(self, content: bytes) -> str:
        """
        Processa CSV e converte para texto estruturado.

        Args:
            content: Conteúdo do CSV

        Returns:
            Texto formatado
        """
        try:
            import csv

            # Decodifica conteúdo
            text = content.decode('utf-8')
            csv_file = io.StringIO(text)
            reader = csv.DictReader(csv_file)

            # Converte para texto estruturado
            text_parts = []
            for row_num, row in enumerate(reader, 1):
                row_text = f"Linha {row_num}:\n"
                for key, value in row.items():
                    if value:
                        row_text += f"  {key}: {value}\n"
                text_parts.append(row_text)

            full_text = "\n".join(text_parts)
            logger.debug(f"Processed CSV: {len(text_parts)} rows")
            return full_text

        except Exception as e:
            logger.error(f"CSV processing error: {e}")
            raise ValueError(f"Failed to process CSV: {e}")

    async def _process_excel(self, content: bytes) -> str:
        """
        Processa Excel e converte para texto.

        Args:
            content: Conteúdo do arquivo Excel

        Returns:
            Texto formatado
        """
        try:
            import openpyxl

            excel_file = io.BytesIO(content)
            workbook = openpyxl.load_workbook(excel_file)

            text_parts = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_parts.append(f"[Planilha: {sheet_name}]")

                for row_num, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    row_values = [str(cell) for cell in row if cell is not None]
                    if row_values:
                        text_parts.append(f"Linha {row_num}: {' | '.join(row_values)}")

                text_parts.append("")  # Linha em branco entre planilhas

            full_text = "\n".join(text_parts)
            logger.debug(f"Processed Excel: {len(workbook.sheetnames)} sheets")
            return full_text

        except Exception as e:
            logger.error(f"Excel processing error: {e}")
            raise ValueError(f"Failed to process Excel: {e}")

    def _chunk_text(self, text: str, chunk_size: int) -> List[str]:
        """
        Divide texto em chunks menores mantendo parágrafos inteiros quando possível.

        Args:
            text: Texto completo
            chunk_size: Tamanho máximo de cada chunk

        Returns:
            Lista de chunks
        """
        if len(text) <= chunk_size:
            return [text]

        chunks = []
        paragraphs = text.split('\n\n')

        current_chunk = ""
        for para in paragraphs:
            # Se parágrafo sozinho já é maior que chunk_size
            if len(para) > chunk_size:
                # Salva chunk atual se houver
                if current_chunk:
                    chunks.append(current_chunk.strip())
                    current_chunk = ""

                # Divide parágrafo por sentença
                sentences = para.split('. ')
                for sentence in sentences:
                    if len(current_chunk) + len(sentence) <= chunk_size:
                        current_chunk += sentence + '. '
                    else:
                        if current_chunk:
                            chunks.append(current_chunk.strip())
                        current_chunk = sentence + '. '
            else:
                # Adiciona parágrafo ao chunk atual
                if len(current_chunk) + len(para) <= chunk_size:
                    current_chunk += para + '\n\n'
                else:
                    if current_chunk:
                        chunks.append(current_chunk.strip())
                    current_chunk = para + '\n\n'

        # Adiciona último chunk
        if current_chunk:
            chunks.append(current_chunk.strip())

        logger.debug(f"Created {len(chunks)} chunks from text")
        return chunks
